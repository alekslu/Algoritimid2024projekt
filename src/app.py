import tkinter as tk
from tkinter import messagebox
from openpyxl import Workbook, load_workbook
from otsingualgoritmid import lineaarotsing, binaarotsing
import os, re, time, json

with open("config.json", "r") as file:
    config = json.load(file)

exceliPath = config["dataFilePath"]
#exceliPath = config["testDataFilePath"]
#exceliPath = config["csvTest"]

class LaohaldusRakendus:
    def __init__(self, root):
        self.root = root
        self.root.title("Laohaldus")
        self.root.geometry("490x505")

        # Load inventory from Excel file
        self.inventory = self.load_inventory_from_excel()

        self.otsi_algorithm = tk.StringVar(value="unselected")

        # Sorteerimise suuna jälgimine igale veerule
        self.sort_directions = {
            "nimetus": False,  # False = kasvav, True = kahanev
            "kategooria": False,
            "kogus": False,
            "hind": False
        }

        self.loo_kasutajaliides()

    def loo_kasutajaliides(self):
        self.toote_nimetus_label = tk.Label(self.root, text="Tootenimetus:")
        self.toote_nimetus_label.grid(row=0, column=0)
        self.toote_nimetus_entry = tk.Entry(self.root)
        self.toote_nimetus_entry.grid(row=0, column=1)

        self.toote_kategooria_label = tk.Label(self.root, text="Kategooria:")
        self.toote_kategooria_label.grid(row=1, column=0)
        self.toote_kategooria_entry = tk.Entry(self.root)
        self.toote_kategooria_entry.grid(row=1, column=1)

        self.toote_kogus_label = tk.Label(self.root, text="Kogus:")
        self.toote_kogus_label.grid(row=2, column=0)
        self.toote_kogus_entry = tk.Entry(self.root)
        self.toote_kogus_entry.grid(row=2, column=1)

        # toode hind Label and Entry
        self.toote_hind_label = tk.Label(self.root, text="Hind:")
        self.toote_hind_label.grid(row=3, column=0)
        self.toote_hind_entry = tk.Entry(self.root)
        self.toote_hind_entry.grid(row=3, column=1)

        # Sortimisvalikud
        self.sort_nimetus_button = tk.Button(self.root, text="Sordi nimetuse järgi", command=lambda: self.sordi_inventory("nimetus"))
        self.sort_kategooria_button = tk.Button(self.root, text="Sordi kategooria järgi", command=lambda: self.sordi_inventory("kategooria"))
        self.sort_kogus_button = tk.Button(self.root, text="Sordi koguse järgi", command=lambda: self.sordi_inventory("kogus"))
        self.sort_hind_button = tk.Button(self.root, text="Sordi hinna järgi", command=lambda: self.sordi_inventory("hind"))

        # Sortimisnuppude paigutused
        self.sort_nimetus_button.grid(row=10, column=0, padx=5, pady=5)
        self.sort_kategooria_button.grid(row=10, column=1, padx=5, pady=5)
        self.sort_kogus_button.grid(row=10, column=2, padx=5, pady=5)
        self.sort_hind_button.grid(row=10, column=3, padx=5, pady=5)

        # lisa toode Button
        self.lisa_toode_button = tk.Button(self.root, text="LISA TOODE", command=self.lisa_toode)
        self.lisa_toode_button.grid(row=4, column=0, columnspan=2)

        # Värskenda Button, algse täisvaate taastamiseks
        self.varskenda_button = tk.Button(self.root, text="VÄRSKENDA", command=self.varskenda_tabel)
        self.varskenda_button.grid(row=7, column=1, columnspan=2)

        # otsi Label and Entry
        self.otsi_label = tk.Label(self.root, text="OTSI toodet:")
        self.otsi_label.grid(row=5, column=0)
        self.otsi_entry = tk.Entry(self.root)
        self.otsi_entry.grid(row=5, column=1)

        # otsi radiobutton  otsingualgoritmi valimiseks
        self.lineaarotsing_rb = tk.Radiobutton(self.root, text="Lineaarotsing", variable=self.otsi_algorithm, value="Linear")
        self.lineaarotsing_rb.grid(row=6, column=0)
        self.binaarotsing_rb = tk.Radiobutton(self.root, text="Binaarotsing", variable=self.otsi_algorithm, value="Binary")
        self.binaarotsing_rb.grid(row=6, column=1)

        # otsi button
        self.otsi_button = tk.Button(self.root, text="OTSI", command=self.otsi_toode)
        self.otsi_button.grid(row=7, column=0, columnspan=3)

        self.results_label = tk.Label(self.root, text="Otsingutulemus: -")
        self.results_label.grid(row=8, column=0, columnspan=3)

        # Tabel
        self.table_frame = tk.Frame(self.root)
        self.table_frame.grid(row=9, column=0, columnspan=3)

        self.table_canvas = tk.Canvas(self.root)
        self.scrollbar = tk.Scrollbar(self.root, orient="vertical", command=self.table_canvas.yview)
        self.table_scroll_frame = tk.Frame(self.table_canvas)

        self.table_canvas.configure(yscrollcommand=self.scrollbar.set)
        self.scrollbar.grid(row=9, column=3, sticky="ns")
        self.table_canvas.grid(row=9, column=0, columnspan=3)
        self.table_window = self.table_canvas.create_window((0, 0), window=self.table_scroll_frame, anchor="nw")
        self.table_scroll_frame.bind("<Configure>", lambda e: self.table_canvas.configure(scrollregion=self.table_canvas.bbox("all")))

        self.v2rskenda_tabel()

    def lisa_toode(self):
        toote_nimetus = self.toote_nimetus_entry.get()
        toote_kategooria = self.toote_kategooria_entry.get()
        toote_kogus = self.toote_kogus_entry.get()
        toote_hind = self.toote_hind_entry.get()

        # Uue toote lisamise valideerimine frontendis
        if not toote_nimetus or not toote_kategooria or not toote_kogus or not toote_hind:
            messagebox.showerror("VIGA!", "Kõik väljad peavad olema täidetud!")
            return

        # Lisatava toote andmed
        toode = {
            "nimetus": toote_nimetus,
            "kategooria": toote_kategooria,
            "kogus": int(toote_kogus),
            "hind": float(toote_hind)
        }
        self.inventory.append(toode)

        # Puhasta väljad kui toode lisatud
        self.toote_nimetus_entry.delete(0, tk.END)
        self.toote_kategooria_entry.delete(0, tk.END)
        self.toote_kogus_entry.delete(0, tk.END)
        self.toote_hind_entry.delete(0, tk.END)

        self.v2rskenda_tabel()

        # Salvesta excelisse
        self.save_inventory_to_excel()

        messagebox.showinfo("Lisamine õnnestus", f"'{toote_nimetus}' lisatud baasi")

    def v2rskenda_tabel(self, inventory=None):
        # Puhasta tabel
        for widget in self.table_scroll_frame.winfo_children():
            widget.destroy()

        # Use the full inventory if no filtered list is provided
        inventory = inventory or self.inventory

        # Päis
        headers = ["Nimetus", "Kategooria", "Kogus", "Hind"]
        for col, header in enumerate(headers):
            header_label = tk.Label(self.table_scroll_frame, text=header, font=('Arial', 10, 'bold'))
            header_label.grid(row=0, column=col, padx=5, pady=5)

        # Tabeli täitmine andmetega
        for row, toode in enumerate(inventory, start=1):
            tk.Label(self.table_scroll_frame, text=toode['nimetus']).grid(row=row, column=0, padx=5, pady=5)
            tk.Label(self.table_scroll_frame, text=toode['kategooria']).grid(row=row, column=1, padx=5, pady=5)
            tk.Label(self.table_scroll_frame, text=toode['kogus']).grid(row=row, column=2, padx=5, pady=5)
            tk.Label(self.table_scroll_frame, text=f"{toode['hind']:,.2f} €").grid(row=row, column=3, padx=5, pady=5)

            # eemalda toode nupp igale tootele
            eemalda_button = tk.Button(self.table_scroll_frame, text="EEMALDA", command=lambda i=toode: self.eemalda_toode(i))
            eemalda_button.grid(row=row, column=4, padx=5, pady=5)

    #Taastab tabeli algse, täisvaate.
    def varskenda_tabel(self):
        self.results_label.config(text="Otsingutulemus: -")
        self.v2rskenda_tabel()

    def eemalda_toode(self, toode):
        self.inventory = [i for i in self.inventory if i != toode]  # eemaldas the selected toode
        self.v2rskenda_tabel()

        # Save updated inventory to Excel
        self.save_inventory_to_excel()

    def otsi_toode(self):
        otsi_term = self.otsi_entry.get().strip()
        if not otsi_term:
            messagebox.showerror("VIGA!", "Sisesta otsingusõna!")
            return

        start_time = time.time()  # Alguse aja salvestamine

        if self.otsi_algorithm.get() == "Linear":
            index = lineaarotsing(self.inventory, otsi_term)
        elif self.otsi_algorithm.get() == "Binary":
            # Binaarotsingu jaoks on oluline, et otsitav väärtus on sorteeritud
            self.inventory.sort(key=lambda x: x['nimetus'].lower())
            index = binaarotsing(self.inventory, otsi_term)
        else:
            messagebox.showerror("VIGA!", "Otsingualgoritm peab olema valitud!")
            return

        end_time = time.time()  # Lõpu aja salvestamine
        duration = (end_time - start_time) * 1000  # Ajakulu millisekundites

        # Kui otsitav leitud, kuva filtreeritud tabel
        if index != -1:
            matched_toode = self.inventory[index]
            filtered_inventory = [matched_toode]
            self.results_label.config(text=f"Leitud: {matched_toode['nimetus']} - {matched_toode['kategooria']} - {matched_toode['kogus']} tk - {matched_toode['hind']} €\n"f"Ajakulu: {duration:.2f} ms")
        else:
            filtered_inventory = []  # Ei leitud vastet
            self.results_label.config(text="Toodet ei leitud!\n"f"Ajakulu: {duration:.2f} ms")

        # Värskenda tabelit, et kuvada ainult otsitud väärtuseid
        self.v2rskenda_tabel(filtered_inventory)


    def load_inventory_from_excel(self):
        file_nimetus = exceliPath
        if not os.path.exists(file_nimetus):
            return []  # Kui faili pole, tagastab tühja listi

        workbook = load_workbook(file_nimetus)
        sheet = workbook.active

        inventory = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            toode = {
                "nimetus": row[0],
                "kategooria": row[1],
                "kogus": row[2],
                "hind": row[3]
            }
            inventory.append(toode)
        return inventory

    def save_inventory_to_excel(self):
        file_nimetus = exceliPath
        workbook = Workbook()
        sheet = workbook.active

        # Create header row
        sheet.append(["Nimetus", "Kategooria", "Kogus", "Hind"])

        # lisa inventory tooted
        for toode in self.inventory:
            sheet.append([toode["nimetus"], toode["kategooria"], toode["kogus"], toode["hind"]])

        workbook.save(file_nimetus)

    # Sordib inventory kindla võtme järgi ja värskendab tabeli. Sortimine käib seljuhul valikuliselt 'nimetus', 'kategooria' või 'hind' järgi.
    def sordi_inventory(self, key):
        # Sorteerimise suuna vahetamine, nupul klikides sorteeritakse loend kahanevas järjestuses ja uuesti klikkides kasvavas järejestuses
        self.sort_directions[key] = not self.sort_directions[key]
        descending = self.sort_directions[key]

        # Rakendame sortimise
        if key in ["nimetus", "kategooria"]:  # Tähestikulised veerud
            if key == "nimetus":
                # Sortimine numbriliselt, kui tootenimedes on numbrid
                self.inventory.sort(key=lambda x: [int(i) if i.isdigit() else i.lower() for i in re.split(r'(\d+)', x[key])], reverse=descending)
            else:
                # Tavaline sortimine kategooria järgi
                self.inventory.sort(key=lambda x: x[key].lower(), reverse=descending)
        elif key in ["kogus", "hind"]:  # Numbrilised veerud
            self.inventory.sort(key=lambda x: x[key], reverse=descending)
        else:
            return

        # Värskendab tabeli sorteeritud inventuuriga
        self.v2rskenda_tabel()

root = tk.Tk()
app = LaohaldusRakendus(root)
root.mainloop()
