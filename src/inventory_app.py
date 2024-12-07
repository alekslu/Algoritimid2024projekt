import tkinter as tk
from tkinter import messagebox
from openpyxl import Workbook, load_workbook
from search_algorithms import lineaarotsing, binaarotsing
from inventory_manager import dataPath
import os

class InventoryApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Inventory Management System")
        self.root.geometry("800x600")  # Set the window size to be larger (800x600)

        # Load inventory from Excel file
        self.inventory = self.load_inventory_from_excel()

        # Create otsi options
        self.otsi_algorithm = tk.StringVar(value="unselected")

        self.create_widgets()

    def create_widgets(self):
        self.toote_nimetus_label = tk.Label(self.root, text="Tootenimetus:")
        self.toote_nimetus_label.grid(row=0, column=0)
        self.toote_nimetus_entry = tk.Entry(self.root)
        self.toote_nimetus_entry.grid(row=0, column=1)

        self.toote_kategooria_label = tk.Label(self.root, text="Kategooria:")
        self.toote_kategooria_label.grid(row=1, column=0)
        self.toote_kategooria_entry = tk.Entry(self.root)
        self.toote_kategooria_entry.grid(row=1, column=1)

        self.toote_kogus_label = tk.Label(self.root, text="Kogus:")
        self.toote_kogus_label.grid(row=2, column=0)
        self.toote_kogus_entry = tk.Entry(self.root)
        self.toote_kogus_entry.grid(row=2, column=1)

        # toode hind Label and Entry
        self.toote_hind_label = tk.Label(self.root, text="Hind:")
        self.toote_hind_label.grid(row=3, column=0)
        self.toote_hind_entry = tk.Entry(self.root)
        self.toote_hind_entry.grid(row=3, column=1)

        # lisa toode Button
        self.lisa_toode_button = tk.Button(self.root, text="LISA TOODE", command=self.lisa_toode)
        self.lisa_toode_button.grid(row=4, column=0, columnspan=2)

        # otsi Label and Entry
        self.otsi_label = tk.Label(self.root, text="OTSI toodet:")
        self.otsi_label.grid(row=5, column=0)
        self.otsi_entry = tk.Entry(self.root)
        self.otsi_entry.grid(row=5, column=1)

        # otsi Radio Buttons for Algorithm Choice
        self.lineaarotsing_rb = tk.Radiobutton(self.root, text="Lineaarotsing", variable=self.otsi_algorithm, value="Linear")
        self.lineaarotsing_rb.grid(row=6, column=0)
        self.binaarotsing_rb = tk.Radiobutton(self.root, text="Binaarotsing", variable=self.otsi_algorithm, value="Binary")
        self.binaarotsing_rb.grid(row=6, column=1)

        # otsi Button
        self.otsi_button = tk.Button(self.root, text="OTSI", command=self.otsi_toode)
        self.otsi_button.grid(row=7, column=0, columnspan=3)

        # Display Results
        self.results_label = tk.Label(self.root, text="Otsingutulemus: -")
        self.results_label.grid(row=8, column=0, columnspan=3)

        # Table to display inventory
        self.table_frame = tk.Frame(self.root)
        self.table_frame.grid(row=9, column=0, columnspan=3)

         # Scrollable Frame for Inventory Table
        self.table_canvas = tk.Canvas(self.root)
        self.scrollbar = tk.Scrollbar(self.root, orient="vertical", command=self.table_canvas.yview)
        self.table_scroll_frame = tk.Frame(self.table_canvas)

        # Configure Scrollbar
        self.table_canvas.configure(yscrollcommand=self.scrollbar.set)
        self.scrollbar.grid(row=9, column=3, sticky="ns")
        self.table_canvas.grid(row=9, column=0, columnspan=3, sticky="nsew")
        
        # Create a window inside the canvas to hold the table frame
        self.table_window = self.table_canvas.create_window((0, 0), window=self.table_scroll_frame, anchor="nw")

        # Bind events to update the canvas size
        self.table_scroll_frame.bind("<Configure>", lambda e: self.table_canvas.configure(scrollregion=self.table_canvas.bbox("all")))

        self.refresh_table()

    def lisa_toode(self):
        toote_nimetus = self.toote_nimetus_entry.get()
        toote_kategooria = self.toote_kategooria_entry.get()
        toote_kogus = self.toote_kogus_entry.get()
        toote_hind = self.toote_hind_entry.get()

        # Validate entries
        if not toote_nimetus or not toote_kategooria or not toote_kogus or not toote_hind:
            messagebox.showerror("VIGA!", "Kõik väljad peavad olema täidetud!")
            return

        # lisa toode to inventory
        toode = {
            "nimetus": toote_nimetus,
            "kategooria": toote_kategooria,
            "kogus": int(toote_kogus),
            "hind": float(toote_hind)
        }
        self.inventory.append(toode)

        # Clear the entry fields
        self.toote_nimetus_entry.delete(0, tk.END)
        self.toote_kategooria_entry.delete(0, tk.END)
        self.toote_kogus_entry.delete(0, tk.END)
        self.toote_hind_entry.delete(0, tk.END)

        self.refresh_table()

        # Save updated inventory to Excel
        self.save_inventory_to_excel()

        messagebox.showinfo("Lisamine õnnestus", f"'{toote_nimetus}' lisatud baasi")

    def refresh_table(self, inventory=None):
        # Clear previous table contents
        for widget in self.table_scroll_frame.winfo_children():
            widget.destroy()

        # Use the full inventory if no filtered list is provided
        inventory = inventory or self.inventory

        # Table headers
        headers = ["Nimetus", "Kategooria", "Kogus", "Hind"]
        for col, header in enumerate(headers):
            header_label = tk.Label(self.table_scroll_frame, text=header, font=('Arial', 10, 'bold'))
            header_label.grid(row=0, column=col, padx=5, pady=5)

        # Populate the table with tooted
        for row, toode in enumerate(inventory, start=1):
            tk.Label(self.table_scroll_frame, text=toode['nimetus']).grid(row=row, column=0, padx=5, pady=5)
            tk.Label(self.table_scroll_frame, text=toode['kategooria']).grid(row=row, column=1, padx=5, pady=5)
            tk.Label(self.table_scroll_frame, text=toode['kogus']).grid(row=row, column=2, padx=5, pady=5)
            tk.Label(self.table_scroll_frame, text=f"{toode['hind']:,.2f} €").grid(row=row, column=3, padx=5, pady=5)

            # eemalda toode button for each row
            eemalda_button = tk.Button(self.table_scroll_frame, text="EEMALDA", command=lambda i=toode: self.eemalda_toode(i))
            eemalda_button.grid(row=row, column=4, padx=5, pady=5)



    def eemalda_toode(self, toode):
        self.inventory = [i for i in self.inventory if i != toode]  # eemaldas the selected toode
        self.refresh_table()

        # Save updated inventory to Excel
        self.save_inventory_to_excel()

    def otsi_toode(self):
        otsi_term = self.otsi_entry.get().strip()
        if not otsi_term:
            messagebox.showerror("VIGA!", "Sisesta otsingusõna!")
            return

        # Perform otsi based on selected algorithm
        if self.otsi_algorithm.get() == "Linear":
            index = lineaarotsing(self.inventory, otsi_term)
        elif self.otsi_algorithm.get() == "Binary":
            # Binary otsi requires data to be sorted first
            self.inventory.sort(key=lambda x: x['nimetus'].lower())
            index = binaarotsing(self.inventory, otsi_term)
        else:
            messagebox.showerror("VIGA!", "Otsingualgoritm peab olema valitud!")
            return

        # If otsi term is found, create a filtered list
        if index != -1:
            matched_toode = self.inventory[index]
            filtered_inventory = [matched_toode]
            self.results_label.config(text=f"Leitud: {matched_toode['nimetus']} - {matched_toode['kategooria']} - {matched_toode['kogus']} tk - {matched_toode['hind']} €")
        else:
            filtered_inventory = []  # No match
            self.results_label.config(text="Toodet ei leitud!")

        # Refresh the table to show only the otsi results
        self.refresh_table(filtered_inventory)


    def load_inventory_from_excel(self):
        file_nimetus = dataPath
        if not os.path.exists(file_nimetus):
            return []  # Return empty list if file doesn't exist

        workbook = load_workbook(file_nimetus)
        sheet = workbook.active

        inventory = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            toode = {
                "nimetus": row[0],
                "kategooria": row[1],
                "kogus": row[2],
                "hind": row[3]
            }
            inventory.append(toode)
        return inventory

    def save_inventory_to_excel(self):
        file_nimetus = dataPath
        workbook = Workbook()
        sheet = workbook.active

        # Create header row
        sheet.append(["Nimetus", "Kategooria", "Kogus", "Hind"])

        # lisa inventory tooted
        for toode in self.inventory:
            sheet.append([toode["nimetus"], toode["kategooria"], toode["kogus"], toode["hind"]])

        workbook.save(file_nimetus)


root = tk.Tk()
app = InventoryApp(root)
root.mainloop()
