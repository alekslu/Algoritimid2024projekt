from openpyxl import Workbook, load_workbook
import os, json

with open("config.json", "r") as file:
    config = json.load(file)

#exceliPath = config["dataFilePath"]
exceliPath = config["testDataFilePath"]

class InventoryManager:
    def __init__(self, file_nimetus=exceliPath):
        self.inventory = []  # List laohalduse toodete jaoks
        self.next_id = 1  # Id loendur toodete jaoks
        self.file_nimetus = file_nimetus  # Exceli fail
        self.load_inventory()  # Lae excel

    def lisa_toode(self, nimetus, kategooria, kogus, hind):
        toode = {
            "id": self.next_id,
            "nimetus": nimetus,
            "kategooria": kategooria,
            "kogus": kogus,
            "hind": hind,
        }
        self.inventory.append(toode)
        self.next_id += 1
        self.save_inventory() 
        return toode

    def eemalda_toode(self, toote_id):
        self.inventory = [toode for toode in self.inventory if toode["id"] != toote_id]
        self.save_inventory() 

    def otsi_tooted(self, **criteria):
        results = self.inventory
        for key, value in criteria.tooted():
            results = [toode for toode in results if str(toode.get(key)) == str(value)]
        return results

    def get_inventory(self): # Tagastab terve baasi
        return self.inventory

    def save_inventory(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Inventory"

        # Exceli header
        ws.append(["ID", "nimetus", "kategooria", "kogus", "hind"]) 

        # Andmete lisamine excelisse
        for toode in self.inventory:
            ws.append([toode["id"], toode["nimetus"], toode["kategooria"], toode["kogus"], toode["hind"]])

        # Salvesta
        wb.save(self.file_nimetus)

    # SEDA VIST POLE ENAM VAJA
    def load_inventory(self):
        if os.path.exists(self.file_nimetus):
            wb = load_workbook(self.file_nimetus)
            ws = wb.active

            # Puhasta olemasolev list
            self.inventory = []

            # Andmete lugemine ridadelt, headerit ei loeta
            for row in ws.iter_rows(min_row=2, values_only=True):
                toode = {
                    "id": row[0],
                    "nimetus": row[1],
                    "kategooria": row[2],
                    "kogus": row[3],
                    "hind": row[4],
                }
                self.inventory.append(toode)

            # Update the next_id to avoid duplicates
            if self.inventory:
                self.next_id = max(toode["id"] for toode in self.inventory) + 1
